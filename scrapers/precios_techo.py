"""Descarga y parseo de la lista oficial de precios techo del Consejo
Nacional de Fijación y Revisión de Precios de Medicamentos de Uso y
Consumo Humano (MSP Ecuador). Actualiza drugs.precio_techo_usd /
fecha_precio_techo. Frecuencia recomendada: semanal (ver docstring de
Boveda Farmacia/Entidades/Secretaría Técnica de Fijación de Precios.md).

Fuente (investigada 2026-08-05): robots.txt de salud.gob.ec permite
/wp-content/uploads/, sin bloqueos. El archivo consolidado es un Excel real
(no escaneado), hoja "ESTRATÉGICOS", encabezado en la fila 10, columnas:
Resolución | Item | Sesión de Consejo | Principio Activo | Primer Nivel de
Desagregación | Forma Farmacéutica (código interno, ej. "SOR-01") |
Concentración | Presentación Comercial | Precio Techo (USD).

El precio techo se fija por segmento de mercado (principio activo +
concentración), casi nunca por marca/laboratorio — la columna
"Presentación Comercial" casi siempre está vacía. Por eso el match contra
nuestro catálogo (`drugs`, poblado por scraping + IA) usa únicamente
principio_activo (contención, ya que el MSP suele incluir la sal química,
ej. "ENALAPRIL MALEATO" vs nuestro "Enalapril") + concentración
(normalizada, sin espacios). La "Forma Farmacéutica" del MSP es un código
interno no comparable directo con nuestros valores canónicos
(tableta/capsula/...), así que no se usa para el match — ver limitación
documentada en la bitácora.

IMPORTANTE: la URL del archivo cambia con cada actualización (incluye la
fecha en el nombre). Este script apunta al vigente al momento de escribirlo
(2026-07-31). Si en una corrida futura falla la descarga, hay que volver a
localizar la URL vigente en
https://www.salud.gob.ec/consejo-nacional-de-fijacion-y-revision-de-precios-de-medicamentos/
antes de asumir que el sitio está caído.

Uso:
    python scrapers/precios_techo.py
    python scrapers/precios_techo.py --dry-run
"""

import argparse
import re
import unicodedata
from datetime import date
from io import BytesIO

import openpyxl

from base import get_html  # noqa: F401  (reservado para localizar la URL vigente automáticamente en el futuro)
from base import BlockedError, USER_AGENT, supabase_patch, supabase_select
import requests

FILE_URL = "https://www.salud.gob.ec/wp-content/uploads/2026/07/Consolidado-precios-techo-actualizado-2026-07-31.xlsx"
FECHA_PUBLICACION = "2026-07-31"

HEADER_ROW = 10
COL_PRINCIPIO_ACTIVO = 4
COL_CONCENTRACION = 7
COL_PRECIO_TECHO = 9


def normalize_text(text):
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    return text.lower().strip()


def normalize_concentracion(text):
    return re.sub(r"\s+", "", normalize_text(text))


def download_workbook():
    resp = requests.get(FILE_URL, headers={"User-Agent": USER_AGENT}, timeout=30)
    if resp.status_code in (403, 429):
        raise BlockedError(f"{FILE_URL} -> {resp.status_code}")
    resp.raise_for_status()
    return openpyxl.load_workbook(BytesIO(resp.content), data_only=True)


def parse_precios_techo():
    wb = download_workbook()
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        principio = row[COL_PRINCIPIO_ACTIVO - 1]
        concentracion = row[COL_CONCENTRACION - 1]
        precio = row[COL_PRECIO_TECHO - 1]
        if not principio or precio is None:
            continue
        try:
            precio = float(precio)
        except (TypeError, ValueError):
            continue
        rows.append(
            {
                "principio_activo_norm": normalize_text(principio),
                "concentracion_norm": normalize_concentracion(concentracion),
                "precio_techo_usd": round(precio, 4),
            }
        )
    return rows


def match_and_update(techo_rows, dry_run=False):
    drugs = supabase_select("drugs", {"select": "id,slug,principio_activo,concentracion"})
    print(f"{len(drugs)} medicamentos en el catálogo maestro, {len(techo_rows)} filas de precio techo descargadas")

    # índice por concentración normalizada -> lista de filas techo, para no
    # comparar cada drug contra las ~1800 filas del Excel una por una.
    by_concentracion = {}
    for t in techo_rows:
        by_concentracion.setdefault(t["concentracion_norm"], []).append(t)

    matched = 0
    ambiguous = 0
    updates = []
    for drug in drugs:
        drug_principio_norm = normalize_text(drug["principio_activo"])
        drug_concentracion_norm = normalize_concentracion(drug["concentracion"])
        candidatos = by_concentracion.get(drug_concentracion_norm, [])
        hits = [t for t in candidatos if drug_principio_norm and drug_principio_norm in t["principio_activo_norm"]]
        if len(hits) == 1:
            matched += 1
            updates.append(
                {
                    "slug": drug["slug"],
                    "precio_techo_usd": hits[0]["precio_techo_usd"],
                    "fecha_precio_techo": FECHA_PUBLICACION,
                }
            )
        elif len(hits) > 1:
            ambiguous += 1  # más de un segmento coincide, no se asigna para no adivinar cuál aplica

    print(f"Emparejados: {matched} | ambiguos (sin asignar): {ambiguous} | sin match: {len(drugs) - matched - ambiguous}")

    if dry_run:
        print("--dry-run: no se escribe a Supabase. Muestra de 5 actualizaciones:")
        for u in updates[:5]:
            print(" ", u)
        return

    for u in updates:
        supabase_patch(
            "drugs",
            {"slug": f"eq.{u['slug']}"},
            {"precio_techo_usd": u["precio_techo_usd"], "fecha_precio_techo": u["fecha_precio_techo"]},
        )
    print(f"{len(updates)} medicamentos actualizados con precio techo.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="No escribir a Supabase")
    args = parser.parse_args()

    try:
        techo_rows = parse_precios_techo()
    except BlockedError as e:
        print(f"BLOQUEADO: {e}. No se evade.")
        return

    match_and_update(techo_rows, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
